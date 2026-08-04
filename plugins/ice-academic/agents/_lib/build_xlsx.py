#!/usr/bin/env python3
"""
build_xlsx.py — XLSX builder ที่บังคับ FONT POLICY + Thai typography discipline อัตโนมัติ
V02R03 | 2026.08.04 | ผูกกับ skill ice-doc-builder §3.0 FONT POLICY + §3.2 E1-E6 + §6 V1-V4

V02R03 — นโยบายย้ายไป `font_policy.py` (SSOT) ไฟล์นี้ import อย่างเดียว ไม่ประกาศซ้ำ
  เหตุผล: RAILS อยู่ในไฟล์นี้ตัวเดียว → build script อีก 5 ตัวใช้ร่วมไม่ได้ → hard-code กันเอง
  ตรวจฟอนต์ทุกฟอร์แมตในที่เดียว → `audit_fonts.py`

V02R02 — ⭐ +V4 RAIL CONFORMANCE (ด่านที่หายไป): V1 ถามแค่ "ชื่อฟอนต์มีจริงไหม" · V2 ถามแค่
"อยู่ blacklist ไหม" → ฟอนต์ที่ **มีจริง + ไม่ blacklist + แต่ผิดนโยบาย** (เช่น Sarabun) ลอดทั้งสองด่าน
เคสจริง 2026.08.04: PWA TCO-Breakdown V01R22 build วันเดียวกับที่นโยบายใช้อยู่แล้ว ยังเป็น Sarabun
และ validator ขึ้น ✅ PASS · ต้นเหตุ: build script เขียนมือตั้ง FONT เองเป็นค่าคงที่ ไม่เคยเรียก RAILS
→ V4 เทียบกับ RAILS[rail] ตรง ๆ · CLI: --rail private|govt · --allow-font "ชื่อ" (TOR บังคับ/ไฟล์รับมา)

V02R02 — แก้ E4 false positive: เดิมฟ้อง "merge ในแถวไทย+wrap" ทุกแถวโดยไม่ดูว่าแถวนั้นตั้ง
row height ไว้แล้วหรือไม่ → ไฟล์ที่ builder ของเราสร้างสด (merge_put ตั้ง height ให้เสมอ) ก็ FAIL
พิสูจน์ด้วย differential test: rebuild จาก script เดิม → FAIL เซลล์ชุดเดียวกันเป๊ะ = ปัญหาอยู่ที่
เครื่องมือตรวจ ไม่ใช่ไฟล์ (หลักการเดียวกับบทเรียน renderer shim 2026.08.01)

ทำไมต้องมีชั้นนี้: กฎที่เป็นตัวหนังสือใน skill ถูกลืมได้ · เคสจริง 2026.07.31 ไฟล์ PWA TOR Matrix
ระบุฟอนต์ "IBM Plex Sans Thai Regular" (ไม่มี family ชื่อนี้จริง) → Excel substitute เงียบ →
ฟอนต์ปน 3 ตัวในไฟล์เดียว · ไม่มี validator ตัวใดในระบบจับได้ → ตอนนี้ V1 จับได้ตั้งแต่ก่อน save

Usage:
    python3 build_xlsx.py spec.json out.xlsx        # ต้องมี marker ICE_BUILD=pipeline นำหน้า
    python3 build_xlsx.py --audit existing.xlsx     # ตรวจไฟล์ที่มีอยู่แล้ว (ไม่เขียนอะไร)

spec.json schema (เพิ่มจาก V01 · ของเดิมใช้ได้เหมือนเดิมทุกอย่าง):
{
  "rail": "private" | "govt",          # เลือกราง (default: private) → กำหนดฟอนต์+ขนาดตาม §3.0
  "font": "…", "font_size": 11,        # override ราง (ใช้เมื่อ TOR/ลูกค้าบังคับ)
  "sheets": [{
      "name": "Sheet1",
      "headers": [...], "rows": [[...]],
      "formulas": {"D2": "=B2+C2"},    # ⚠ ระวัง §2B.3 Removed Records
      "freeze": "B2",
      "column_widths": {"A": 30},
      "header_fill": "1F4E79", "header_font_color": "FFFFFF",
      "wrap_columns": ["C","G"],       # คอลัมน์ที่ให้ wrap (จะคำนวณความสูงแถวให้)
      "row_height_lines": {"5": 2}     # override จำนวนบรรทัดที่คาดของแถวนั้น
  }]
}
"""
import sys, os, json, glob, re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─────────────────────────────────────────────────────────────────────────────
# §3.0 FONT POLICY — ⭐ import จาก SSOT (V02R03) · ห้าม re-declare ที่ไฟล์นี้อีก
#   บทเรียน 2026.08.04: RAILS เคยถูกประกาศซ้ำในไฟล์นี้ → build script ฟอร์แมตอื่น
#   (pptx/docx/dashboard/deck/html) ไม่มีทางใช้ร่วมได้ → hard-code กันเอง 5 ระบบ
#   → นโยบายที่ LOCKED ไว้ บังคับใช้ไม่ได้จริงทั้งระบบ (เคส PWA TCO V01R22)
# ─────────────────────────────────────────────────────────────────────────────
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from font_policy import (RAILS, BLACKLIST_PATTERNS, LATIN_ONLY, THAI_RE,   # noqa: E402
                         APPROVED_ALT, RETIRED,
                         has_thai, installed_families, blacklist_hit, check_fonts)

ROW_H_FACTOR = 1.45   # §3.2 E2 — ทดสอบแล้ว
ROW_H_PAD = 6


# ─────────────────────────────────────────────────────────────────────────────
# VALIDATOR — §6 V1/V2 + §3.2 E2-E5
# ─────────────────────────────────────────────────────────────────────────────
def audit(path: str, strict: bool = True, rail: str = "private",
          allow_fonts=None) -> dict:   # ⚠ ไม่ใช้ `set | None` — python3 บนเครื่องนี้คือ 3.9
    """rail = private|govt (ตาม §3.0) · allow_fonts = ฟอนต์ที่ยกเว้นได้ (TOR บังคับ / ไฟล์ที่ได้รับมา)"""
    allow_fonts = allow_fonts or set()
    wb = load_workbook(path)
    fams = installed_families()
    rep = {"fonts_used": set(), "unresolvable": [], "blacklisted": [],
           "thai_cells": 0, "thai_latin_only_font": [], "thai_no_rowheight": [],
           "merged_thai_rows": [], "merged_thai_rows_ok": [],
           "shrink_to_fit": [], "formula_cells": 0,
           "v1_skipped": not bool(fams)}

    for ws in wb.worksheets:
        merged_rows = set()
        for rng in ws.merged_cells.ranges:
            merged_rows.update(range(rng.min_row, rng.max_row + 1))
        for row in ws.iter_rows():
            for c in row:
                if c.value is None:
                    continue
                fname = c.font.name if c.font else None
                if fname:
                    rep["fonts_used"].add(fname)
                if isinstance(c.value, str) and c.value.startswith("="):
                    rep["formula_cells"] += 1
                if not has_thai(c.value):
                    continue
                rep["thai_cells"] += 1
                loc = f"{ws.title}!{c.coordinate}"
                if not fname or fname in LATIN_ONLY:
                    rep["thai_latin_only_font"].append((loc, fname or "(ไม่ระบุ)"))
                if c.alignment and c.alignment.shrinkToFit:
                    rep["shrink_to_fit"].append(loc)
                wrapped = bool(c.alignment and c.alignment.wrap_text)
                if wrapped:
                    has_h = ws.row_dimensions[c.row].height is not None
                    if not has_h:
                        rep["thai_no_rowheight"].append(loc)
                    # E4 — merge ฆ่า AutoFit จริง แต่เป็น "ข้อบกพร่อง" ก็ต่อเมื่อแถวนั้น *ต้องพึ่ง* AutoFit
                    #   ตั้ง row height ไว้ชัดเจนแล้ว = escape hatch ที่ builder เราใช้ตั้งใจ (merge_put) → ผ่าน
                    #   V02R02 (บทเรียน PWA 2026.08.04): เดิมไม่เช็คเงื่อนไขนี้ → ไฟล์ที่ script สร้างสด
                    #   และผ่าน assert ตัวเอง ก็ยัง FAIL E4 = false positive ทุกแถว banner/band
                    if c.row in merged_rows:
                        (rep["merged_thai_rows"] if not has_h
                         else rep["merged_thai_rows_ok"]).append(loc)

    for name in sorted(rep["fonts_used"]):
        why = blacklist_hit(name)
        if why:
            rep["blacklisted"].append((name, why))
        if fams and name not in fams and name not in LATIN_ONLY:
            rep["unresolvable"].append(name)

    # §3.5 T2 — ทำนายจุดตัดกลางคำ (เตือน ไม่ fail: แก้ที่ความกว้างคอลัมน์ ไม่ใช่ที่ข้อความ)
    try:
        sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
        from thai_wordbreak import find_bad_breaks
        risky = 0
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    if has_thai(c.value) and c.alignment and c.alignment.wrap_text:
                        w = ws.column_dimensions[c.column_letter].width or 20
                        if find_bad_breaks(c.value, w):
                            risky += 1
        rep["thai_midword_break"] = risky
        if risky:
            print(f"⚠ §3.5 T2: {risky} เซลล์เสี่ยงตัดบรรทัดกลางคำ → "
                  f"`python3 _lib/thai_wordbreak.py --audit <file>` ดูรายจุด "
                  f"(แก้: ①ขยายคอลัมน์ ②ปรับข้อความ ③ZWSP-แลกกับ Ctrl+F)")
    except Exception:
        pass

    # ⭐ V02R04: V1/V2/V4/V5 เรียก check_fonts จาก SSOT — เลิกมีสำเนากฎในไฟล์นี้
    #   (เดิม V02R02 เขียน logic ซ้ำที่นี่ = ตรงข้ามกับเหตุผลที่แยก font_policy ออกมาแต่แรก)
    fc = check_fonts(rep["fonts_used"], rail=rail, allow_fonts=allow_fonts, fams=fams)
    for k in ("off_rail", "retired", "alt_used", "rail"):
        rep[k] = fc[k]
    fails = list(fc["fails"])

    if rep["thai_latin_only_font"]:
        fails.append(f"เซลล์ไทยได้ฟอนต์ที่ไม่มี glyph ไทย: {rep['thai_latin_only_font'][:5]}")
    if rep["merged_thai_rows"]:
        fails.append(f"E4 merge ในแถวไทย+wrap *และไม่ตั้ง row height* (AutoFit ตาย ไม่มีอะไรกู้ได้): "
                     f"{rep['merged_thai_rows'][:5]}")
    if rep["shrink_to_fit"]:
        fails.append(f"E5 shrink-to-fit บนเซลล์ไทย: {rep['shrink_to_fit'][:5]}")
    if rep["thai_no_rowheight"]:
        fails.append(f"E2 เซลล์ไทย+wrap ไม่ตั้ง row height: {rep['thai_no_rowheight'][:5]}")
    rep["fails"] = fails

    print(f"VALIDATOR | thai_cells={rep['thai_cells']} · fonts={len(rep['fonts_used'])} "
          f"· formulas={rep['formula_cells']} · V1={'SKIPPED(no fontTools)' if rep['v1_skipped'] else 'ran'}")
    print(f"  fonts: {', '.join(sorted(rep['fonts_used'])) or '(none)'}")
    if rep.get("alt_used"):
        print(f"  ℹ ใช้ตัวเลือกอนุมัติ (ไม่ใช่ฟอนต์ราง): {', '.join(rep['alt_used'])} "
              f"— ผ่านได้ แต่ GAP ไทย-ละตินกว้างกว่า ตรวจว่าไทยไม่ดูเล็กเกินไป")
    if rep["merged_thai_rows_ok"]:
        # แสดงเสมอ — ไม่เงียบ: ผู้อ่านต้องรู้ว่าเรา "ยกเว้น" อะไรไป ไม่ใช่ "ไม่เจอ"
        print(f"  ℹ E4 ยกเว้น {len(rep['merged_thai_rows_ok'])} เซลล์ merge+wrap ที่ตั้ง row height ไว้แล้ว "
              f"(escape hatch ถูกกฎ) เช่น {rep['merged_thai_rows_ok'][:3]}")
    if fails:
        print("❌ FAIL:")
        for f in fails:
            print("   -", f)
        if strict:
            sys.exit(3)
    else:
        print(f"✅ PASS — V1 resolve ครบ · V2 ไม่มี blacklist · V4 ตรงราง '{rail}' · E2/E4/E5 ผ่าน")
    return rep


# ─────────────────────────────────────────────────────────────────────────────
# BUILD
# ─────────────────────────────────────────────────────────────────────────────
def build(spec_path: str, out_path: str):
    with open(spec_path) as f:
        spec = json.load(f)

    rail = spec.get("rail", "private")
    if rail not in RAILS:
        sys.exit(f"rail ต้องเป็น private|govt (ได้: {rail})")
    FONT = spec.get("font", RAILS[rail]["font"])
    SIZE = spec.get("font_size", RAILS[rail]["size"])

    # V1/V2 ตรวจ "ก่อน" build — ล้มเร็วดีกว่าล้มช้า
    why = blacklist_hit(FONT)
    if why:
        sys.exit(f"❌ V2 BLACKLIST: '{FONT}' — {why}")
    fams = installed_families()
    if fams and FONT not in fams:
        near = [f for f in fams if FONT.split()[0].lower() in f.lower()][:6]
        sys.exit(f"❌ V1 FONT-NAME ไม่ resolve: '{FONT}' ไม่มีใน name table ของฟอนต์ที่ติดตั้ง\n"
                 f"   ชื่อใกล้เคียงที่มีจริง: {near}\n"
                 f"   (กับดักที่พบบ่อย: เติม subfamily ต่อท้าย เช่น '... Regular')")

    wb = Workbook()
    wb.remove(wb.active)

    # §3.2 E6 — ตั้ง default font ก่อน แล้วค่อยคำนวณ column width
    for ns in wb._named_styles:
        if ns.name == "Normal":
            ns.font = Font(name=FONT, size=SIZE)

    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for sht in spec.get("sheets", []):
        ws = wb.create_sheet(title=sht["name"][:31])
        wrap_cols = set(sht.get("wrap_columns") or [])
        headers = sht.get("headers", [])
        if headers:
            ws.append(headers)
            hf = PatternFill("solid", fgColor=sht.get("header_fill", "1F4E79"))
            hfont = Font(name=FONT, size=SIZE, bold=True,
                         color=sht.get("header_font_color", "FFFFFF"))
            for j, _ in enumerate(headers, start=1):
                c = ws.cell(row=1, column=j)
                c.fill, c.font = hf, hfont
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in sht.get("rows", []):
            ws.append(row)
        for cell, formula in (sht.get("formulas") or {}).items():
            ws[cell] = formula

        overrides = {int(k): v for k, v in (sht.get("row_height_lines") or {}).items()}

        for row in ws.iter_rows():
            thai_in_row = False
            wrapped_in_row = False
            for c in row:
                col = c.column_letter
                is_thai = has_thai(c.value)
                thai_in_row = thai_in_row or is_thai
                # ⭐ ฟอนต์เดียวทุกเซลล์ (§3.0 single-family) — กัน default leak
                keep_bold = bool(c.font and c.font.bold)
                keep_color = c.font.color if (c.font and c.font.color) else None
                c.font = Font(name=FONT, size=SIZE, bold=keep_bold, color=keep_color)
                c.border = border
                want_wrap = col in wrap_cols
                wrapped_in_row = wrapped_in_row or want_wrap
                # §3.2 E3 vertical=center · E5 ไม่ตั้ง shrinkToFit เด็ดขาด
                c.alignment = Alignment(
                    horizontal=(c.alignment.horizontal if c.alignment else None),
                    vertical="center", wrap_text=want_wrap or (c.alignment.wrap_text if c.alignment else False),
                    shrink_to_fit=False,
                )
            # §3.2 E2 — ตั้ง row height ชัดเจนเมื่อมีไทย (หรือมี wrap)
            if thai_in_row or wrapped_in_row:
                lines = overrides.get(row[0].row)
                if lines is None:
                    longest = max((len(str(c.value)) for c in row if c.value), default=0)
                    lines = 2 if (wrapped_in_row and longest > 60) else 1
                ws.row_dimensions[row[0].row].height = round(SIZE * ROW_H_FACTOR * lines + ROW_H_PAD, 1)

        for col, width in (sht.get("column_widths") or {}).items():
            ws.column_dimensions[col].width = width
        if sht.get("freeze"):
            ws.freeze_panes = sht["freeze"]

    wb.save(out_path)
    if not os.path.exists(out_path):          # กฎเหล็ก: tool สำเร็จ ≠ ไฟล์เกิด
        sys.exit("❌ save รายงานสำเร็จแต่ไม่พบไฟล์")
    print(f"OK: {out_path} · {len(wb.sheetnames)} sheets · rail={rail} · font={FONT} {SIZE}pt")
    print(f"⚠ §3.2 E1: Excel ฝังฟอนต์ไม่ได้ — ส่งลูกค้าต้องแนบ PDF companion หรือใช้ฟอนต์ที่ลูกค้ามี "
          f"(fallback: {RAILS[rail]['fallback']})")
    # ⭐ V02R02: ต้องส่ง rail ที่ build ใช้จริง — ไม่งั้น build รางราชการจะ FAIL V4 กับ default private
    audit(out_path, rail=rail, allow_fonts={FONT})


if __name__ == "__main__":
    argv = sys.argv[1:]
    rail, allow = "private", set()
    while len(argv) >= 2 and argv[0] in ("--rail", "--allow-font"):
        if argv[0] == "--rail":
            rail = argv[1]
            if rail not in RAILS:
                sys.exit(f"--rail ต้องเป็น private|govt (ได้: {rail})")
        else:
            allow.add(argv[1])
        argv = argv[2:]

    if len(argv) == 2 and argv[0] == "--audit":
        audit(argv[1], rail=rail, allow_fonts=allow); sys.exit(0)
    if len(argv) != 2:
        print("usage: build_xlsx.py [--rail private|govt] [--allow-font NAME]... "
              "(spec.json out.xlsx | --audit file.xlsx)", file=sys.stderr)
        sys.exit(2)
    build(argv[0], argv[1])
