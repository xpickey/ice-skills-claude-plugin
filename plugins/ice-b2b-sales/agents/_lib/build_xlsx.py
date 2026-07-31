#!/usr/bin/env python3
"""
build_xlsx.py — XLSX builder ที่บังคับ FONT POLICY + Thai typography discipline อัตโนมัติ
V02R01 | 2026.07.31 | ผูกกับ skill ice-doc-builder §3.0 FONT POLICY + §3.2 E1-E6 + §6 V1-V3

ทำไมต้องมีชั้นนี้: กฎที่เป็นตัวหนังสือใน skill ถูกลืมได้ · เคสจริง 2026.07.31 ไฟล์ PWA TOR Matrix
ระบุฟอนต์ "IBM Plex Sans Thai Regular" (ไม่มี family ชื่อนี้จริง) → Excel substitute เงียบ →
ฟอนต์ปน 3 ตัวในไฟล์เดียว · ไม่มี validator ตัวใดในระบบจับได้ → ตอนนี้ V1 จับได้ตั้งแต่ก่อน save

Usage:
    python3 build_xlsx.py spec.json out.xlsx        # ต้องมี marker ICE_BUILD=pipeline นำหน้า
    python3 build_xlsx.py --audit existing.xlsx     # ตรวจไฟล์ที่มีอยู่แล้ว (ไม่เขียนอะไร)

spec.json schema (เพิ่มจาก V01 · ของเดิมใช้ได้เหมือนเดิมทุกอย่าง):
{
  "rail": "private" | "govt",          # เลือกราง (default: private) → กำหนดฟอนต์+ขนาดตาม §3.0
  "font": "…", "font_size": 11,        # override ราง (ใช้เมื่อ TOR/ลูกค้าบังคับ)
  "sheets": [{
      "name": "Sheet1",
      "headers": [...], "rows": [[...]],
      "formulas": {"D2": "=B2+C2"},    # ⚠ ระวัง §2B.3 Removed Records
      "freeze": "B2",
      "column_widths": {"A": 30},
      "header_fill": "1F4E79", "header_font_color": "FFFFFF",
      "wrap_columns": ["C","G"],       # คอลัมน์ที่ให้ wrap (จะคำนวณความสูงแถวให้)
      "row_height_lines": {"5": 2}     # override จำนวนบรรทัดที่คาดของแถวนั้น
  }]
}
"""
import sys, os, json, glob, re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─────────────────────────────────────────────────────────────────────────────
# §3.0 FONT POLICY — 2 ราง (LOCKED โดย user 2026.07.31)
# ─────────────────────────────────────────────────────────────────────────────
RAILS = {
    # เอกชน: ไทย=อังกฤษ ไม่บวก pt (cap 0.698 em — ละตินอยู่ในตัวเดียวกัน)
    "private": {"font": "IBM Plex Sans Thai Looped", "size": 11, "fallback": "Tahoma"},
    # ราชการ/TOR/e-GP: 16pt = ละติน 11-12pt (วัดได้ ×1.47)
    "govt":    {"font": "TH Sarabun New",            "size": 16, "fallback": "Tahoma"},
}

# §3.0 BLACKLIST — เหตุผลรายตัวอยู่ใน skill
BLACKLIST_PATTERNS = [
    (r"^TH Sarabun ?IT", "แปลงเลขอารบิก→เลขไทยเงียบ ๆ + ชื่อชนกับ PSK + digit width +24%"),
    (r"^Angsana",   "ทำลาย สระอำ ในชั้นข้อความ 100% (copy-paste/ค้นหา/index พัง)"),
    (r"^Cordia",    "ทำลาย สระอำ 100% · ไม่มีบน macOS"),
    (r"^Browallia", "ทำลาย สระอำ 100% · ไม่มีบน macOS"),
    (r"^Eucrosia",  "ตระกูล UPC เดียวกัน"),
    (r"^Jasmine",   "ตระกูล UPC เดียวกัน · ไม่ได้ติดตั้ง"),
    (r"^Microsoft Sans Serif", "ไม่มี Bold จริง + ที่ว่างวรรณยุกต์ = 0"),
]
LATIN_ONLY = {"Calibri", "Aptos", "Arial", "Cambria", "Times New Roman", "Helvetica"}

THAI_RE = re.compile(r"[฀-๿]")
ROW_H_FACTOR = 1.45   # §3.2 E2 — ทดสอบแล้ว
ROW_H_PAD = 6


def has_thai(v) -> bool:
    return isinstance(v, str) and bool(THAI_RE.search(v))


def installed_families() -> set:
    """§6 V1 — family name จริงจาก name table (nameID 1) ของฟอนต์ที่ติดตั้ง"""
    fams = set()
    try:
        from fontTools.ttLib import TTFont
    except ImportError:
        return fams          # ไม่มี fontTools → V1 ข้าม (รายงานว่าข้าม ไม่ใช่ผ่าน)
    roots = ["/System/Library/Fonts", "/System/Library/Fonts/Supplemental",
             "/Library/Fonts", os.path.expanduser("~/Library/Fonts")]
    for root in roots:
        for ext in ("ttf", "otf", "ttc"):
            for p in glob.glob(f"{root}/**/*.{ext}", recursive=True):
                try:
                    f = TTFont(p, fontNumber=0, lazy=True)
                    n = f["name"].getDebugName(1)
                    if n:
                        fams.add(n)
                except Exception:
                    pass
    return fams


def blacklist_hit(name: str):
    for pat, reason in BLACKLIST_PATTERNS:
        if re.match(pat, name or "", re.I):
            return reason
    return None


# ─────────────────────────────────────────────────────────────────────────────
# VALIDATOR — §6 V1/V2 + §3.2 E2-E5
# ─────────────────────────────────────────────────────────────────────────────
def audit(path: str, strict: bool = True) -> dict:
    wb = load_workbook(path)
    fams = installed_families()
    rep = {"fonts_used": set(), "unresolvable": [], "blacklisted": [],
           "thai_cells": 0, "thai_latin_only_font": [], "thai_no_rowheight": [],
           "merged_thai_rows": [], "shrink_to_fit": [], "formula_cells": 0,
           "v1_skipped": not bool(fams)}

    for ws in wb.worksheets:
        merged_rows = set()
        for rng in ws.merged_cells.ranges:
            merged_rows.update(range(rng.min_row, rng.max_row + 1))
        for row in ws.iter_rows():
            for c in row:
                if c.value is None:
                    continue
                fname = c.font.name if c.font else None
                if fname:
                    rep["fonts_used"].add(fname)
                if isinstance(c.value, str) and c.value.startswith("="):
                    rep["formula_cells"] += 1
                if not has_thai(c.value):
                    continue
                rep["thai_cells"] += 1
                loc = f"{ws.title}!{c.coordinate}"
                if not fname or fname in LATIN_ONLY:
                    rep["thai_latin_only_font"].append((loc, fname or "(ไม่ระบุ)"))
                if c.alignment and c.alignment.shrinkToFit:
                    rep["shrink_to_fit"].append(loc)
                wrapped = bool(c.alignment and c.alignment.wrap_text)
                if wrapped:
                    if ws.row_dimensions[c.row].height is None:
                        rep["thai_no_rowheight"].append(loc)
                    if c.row in merged_rows:
                        rep["merged_thai_rows"].append(loc)

    for name in sorted(rep["fonts_used"]):
        why = blacklist_hit(name)
        if why:
            rep["blacklisted"].append((name, why))
        if fams and name not in fams and name not in LATIN_ONLY:
            rep["unresolvable"].append(name)

    fails = []
    if rep["unresolvable"]:
        fails.append(f"V1 FONT-NAME ไม่ resolve: {rep['unresolvable']}")
    if rep["blacklisted"]:
        fails.append(f"V2 BLACKLIST: {[n for n, _ in rep['blacklisted']]}")
    if rep["thai_latin_only_font"]:
        fails.append(f"เซลล์ไทยได้ฟอนต์ที่ไม่มี glyph ไทย: {rep['thai_latin_only_font'][:5]}")
    if rep["merged_thai_rows"]:
        fails.append(f"E4 merge ในแถวไทย+wrap (AutoFit ตาย): {rep['merged_thai_rows'][:5]}")
    if rep["shrink_to_fit"]:
        fails.append(f"E5 shrink-to-fit บนเซลล์ไทย: {rep['shrink_to_fit'][:5]}")
    if rep["thai_no_rowheight"]:
        fails.append(f"E2 เซลล์ไทย+wrap ไม่ตั้ง row height: {rep['thai_no_rowheight'][:5]}")
    rep["fails"] = fails

    print(f"VALIDATOR | thai_cells={rep['thai_cells']} · fonts={len(rep['fonts_used'])} "
          f"· formulas={rep['formula_cells']} · V1={'SKIPPED(no fontTools)' if rep['v1_skipped'] else 'ran'}")
    print(f"  fonts: {', '.join(sorted(rep['fonts_used'])) or '(none)'}")
    if fails:
        print("❌ FAIL:")
        for f in fails:
            print("   -", f)
        if strict:
            sys.exit(3)
    else:
        print("✅ PASS — V1 resolve ครบ · V2 ไม่มี blacklist · E2/E4/E5 ผ่าน")
    return rep


# ─────────────────────────────────────────────────────────────────────────────
# BUILD
# ─────────────────────────────────────────────────────────────────────────────
def build(spec_path: str, out_path: str):
    with open(spec_path) as f:
        spec = json.load(f)

    rail = spec.get("rail", "private")
    if rail not in RAILS:
        sys.exit(f"rail ต้องเป็น private|govt (ได้: {rail})")
    FONT = spec.get("font", RAILS[rail]["font"])
    SIZE = spec.get("font_size", RAILS[rail]["size"])

    # V1/V2 ตรวจ "ก่อน" build — ล้มเร็วดีกว่าล้มช้า
    why = blacklist_hit(FONT)
    if why:
        sys.exit(f"❌ V2 BLACKLIST: '{FONT}' — {why}")
    fams = installed_families()
    if fams and FONT not in fams:
        near = [f for f in fams if FONT.split()[0].lower() in f.lower()][:6]
        sys.exit(f"❌ V1 FONT-NAME ไม่ resolve: '{FONT}' ไม่มีใน name table ของฟอนต์ที่ติดตั้ง\n"
                 f"   ชื่อใกล้เคียงที่มีจริง: {near}\n"
                 f"   (กับดักที่พบบ่อย: เติม subfamily ต่อท้าย เช่น '... Regular')")

    wb = Workbook()
    wb.remove(wb.active)

    # §3.2 E6 — ตั้ง default font ก่อน แล้วค่อยคำนวณ column width
    for ns in wb._named_styles:
        if ns.name == "Normal":
            ns.font = Font(name=FONT, size=SIZE)

    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for sht in spec.get("sheets", []):
        ws = wb.create_sheet(title=sht["name"][:31])
        wrap_cols = set(sht.get("wrap_columns") or [])
        headers = sht.get("headers", [])
        if headers:
            ws.append(headers)
            hf = PatternFill("solid", fgColor=sht.get("header_fill", "1F4E79"))
            hfont = Font(name=FONT, size=SIZE, bold=True,
                         color=sht.get("header_font_color", "FFFFFF"))
            for j, _ in enumerate(headers, start=1):
                c = ws.cell(row=1, column=j)
                c.fill, c.font = hf, hfont
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in sht.get("rows", []):
            ws.append(row)
        for cell, formula in (sht.get("formulas") or {}).items():
            ws[cell] = formula

        overrides = {int(k): v for k, v in (sht.get("row_height_lines") or {}).items()}

        for row in ws.iter_rows():
            thai_in_row = False
            wrapped_in_row = False
            for c in row:
                col = c.column_letter
                is_thai = has_thai(c.value)
                thai_in_row = thai_in_row or is_thai
                # ⭐ ฟอนต์เดียวทุกเซลล์ (§3.0 single-family) — กัน default leak
                keep_bold = bool(c.font and c.font.bold)
                keep_color = c.font.color if (c.font and c.font.color) else None
                c.font = Font(name=FONT, size=SIZE, bold=keep_bold, color=keep_color)
                c.border = border
                want_wrap = col in wrap_cols
                wrapped_in_row = wrapped_in_row or want_wrap
                # §3.2 E3 vertical=center · E5 ไม่ตั้ง shrinkToFit เด็ดขาด
                c.alignment = Alignment(
                    horizontal=(c.alignment.horizontal if c.alignment else None),
                    vertical="center", wrap_text=want_wrap or (c.alignment.wrap_text if c.alignment else False),
                    shrink_to_fit=False,
                )
            # §3.2 E2 — ตั้ง row height ชัดเจนเมื่อมีไทย (หรือมี wrap)
            if thai_in_row or wrapped_in_row:
                lines = overrides.get(row[0].row)
                if lines is None:
                    longest = max((len(str(c.value)) for c in row if c.value), default=0)
                    lines = 2 if (wrapped_in_row and longest > 60) else 1
                ws.row_dimensions[row[0].row].height = round(SIZE * ROW_H_FACTOR * lines + ROW_H_PAD, 1)

        for col, width in (sht.get("column_widths") or {}).items():
            ws.column_dimensions[col].width = width
        if sht.get("freeze"):
            ws.freeze_panes = sht["freeze"]

    wb.save(out_path)
    if not os.path.exists(out_path):          # กฎเหล็ก: tool สำเร็จ ≠ ไฟล์เกิด
        sys.exit("❌ save รายงานสำเร็จแต่ไม่พบไฟล์")
    print(f"OK: {out_path} · {len(wb.sheetnames)} sheets · rail={rail} · font={FONT} {SIZE}pt")
    print(f"⚠ §3.2 E1: Excel ฝังฟอนต์ไม่ได้ — ส่งลูกค้าต้องแนบ PDF companion หรือใช้ฟอนต์ที่ลูกค้ามี "
          f"(fallback: {RAILS[rail]['fallback']})")
    audit(out_path)


if __name__ == "__main__":
    if len(sys.argv) == 3 and sys.argv[1] == "--audit":
        audit(sys.argv[2]); sys.exit(0)
    if len(sys.argv) != 3:
        print("usage: build_xlsx.py spec.json out.xlsx | build_xlsx.py --audit file.xlsx",
              file=sys.stderr)
        sys.exit(2)
    build(sys.argv[1], sys.argv[2])
