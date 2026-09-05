#!/usr/bin/env python3
"""
thai_wordbreak.py — ตัดคำไทยให้บรรทัดสวย ใน .docx / .xlsx / .pptx
V01R02 | 2026.08.30 | ผูกกับ skill ice-doc-builder §3.5
(V01R02: --audit แสดง 12 รายการแรกพร้อมบรรทัดแจ้งยอดที่เหลือชัดเจน + เพิ่ม --all ดูครบทุกรายการ
 — แก้เคสรายงาน 12 ทั้งที่เสี่ยงจริง 77 · บทเรียน 70-Fleet/thai-wordbreak-audit-truncated-output.md)

ปัญหา: ไทยไม่มีช่องว่างระหว่างคำ → Office ตัดบรรทัดกลางคำ ("ระบบบัญ / ชีแยกประเภท")
เครื่องมือ: PyThaiNLP (newmm engine — รักษา case ของอังกฤษ ต่างจาก longest ที่ lowercase ทิ้ง)

⭐ 3 ชั้น เรียงจาก "ไม่แตะข้อความ" ไป "แตะข้อความ" — ใช้ชั้นบนก่อนเสมอ
   T1 LANG-TAG   : ตั้ง lang="th-TH" บน run → ให้ engine ของ Office ตัดคำเอง
                   ✅ ข้อความไม่ถูกแก้ · Ctrl+F ยังหาเจอ · copy-paste สะอาด
                   ใช้ได้: Word (w:lang w:bidi) · PowerPoint (a:rPr lang) — ไม่มีใน cell ของ Excel
   T2 QA-ONLY    : ใช้ PyThaiNLP "ทำนาย" ว่าบรรทัดจะตัดกลางคำตรงไหน → เตือน/ขยายคอลัมน์แทน
                   ✅ ข้อความไม่ถูกแก้เลย · เป็น validator ไม่ใช่ transformer
   T3 ZWSP       : แทรก U+200B ที่ขอบคำ → บังคับจุดตัดบรรทัด
                   ⚠ ราคา: Ctrl+F หาคำที่คร่อม ZWSP ไม่เจอ · copy ได้อักขระซ่อนติดไป
                   ใช้เมื่อ: Excel cell (ไม่มี T1) หรือ T1 ให้ผลไม่ดีพอ และ layout สำคัญกว่า search

Usage:
    python3 thai_wordbreak.py --check "ข้อความ" --width 30    # T2 ทำนายจุดตัด
    python3 thai_wordbreak.py --zwsp  "ข้อความ"               # T3 ใส่ ZWSP
    python3 thai_wordbreak.py --strip "ข้อความ"               # ถอด ZWSP ออก
    python3 thai_wordbreak.py --audit file.xlsx --col C --width 45   # แสดง 12 รายการแรก
    python3 thai_wordbreak.py --audit file.xlsx --all                # แสดงครบทุกรายการ
"""
import sys, re, unicodedata

ZWSP = "​"
THAI_RE = re.compile(r"[฀-๿]")
# อักขระไทยที่ "ไม่กินความกว้าง" (สระบน/ล่าง + วรรณยุกต์) — ต้องไม่นับตอนวัดความยาวบรรทัด
COMBINING = set("ั") | set(chr(c) for c in range(0x0E34, 0x0E3B)) | \
            set(chr(c) for c in range(0x0E47, 0x0E4F))

_ENGINE = "newmm"   # ห้ามใช้ longest — มัน lowercase อังกฤษทิ้ง (SAP → sap)


def has_thai(s) -> bool:
    return isinstance(s, str) and bool(THAI_RE.search(s))


def display_width(s: str) -> float:
    """ความกว้างโดยประมาณเป็น 'ตัวอักษร' — combining mark นับ 0"""
    return sum(0 if ch in COMBINING else 1 for ch in s)


def segment(text: str, engine: str = _ENGINE) -> list:
    """ตัดคำด้วย PyThaiNLP · ไม่มี pythainlp → คืนทั้งก้อน (fail-soft พร้อมเตือน)"""
    if not has_thai(text):
        return [text]
    try:
        from pythainlp.tokenize import word_tokenize
    except ImportError:
        print("WARN: ไม่มี pythainlp — ข้ามการตัดคำ (pip install pythainlp)", file=sys.stderr)
        return [text]
    return word_tokenize(text, engine=engine, keep_whitespace=True)


def zwsp(text: str, engine: str = _ENGINE) -> str:
    """T3 — แทรก ZWSP ที่ขอบคำไทย · ไม่แทรกซ้ำ ไม่แทรกติดช่องว่างที่มีอยู่"""
    if not has_thai(text):
        return text
    toks = segment(text, engine)
    out = []
    for i, t in enumerate(toks):
        if i and t and out and out[-1]:
            prev = out[-1][-1]
            # ข้ามถ้ามีตัวคั่นอยู่แล้ว หรือฝั่งใดไม่ใช่ไทย (ไม่ยุ่งกับคำอังกฤษ/ตัวเลข)
            if not (prev.isspace() or t[0].isspace() or prev == ZWSP) \
               and (has_thai(prev) or has_thai(t[0])):
                out.append(ZWSP)
        out.append(t)
    return "".join(out)


def strip_zwsp(text: str) -> str:
    return text.replace(ZWSP, "") if isinstance(text, str) else text


def find_bad_breaks(text: str, width: float, engine: str = _ENGINE) -> list:
    """T2 — ทำนายว่าถ้า wrap ที่ความกว้างนี้ บรรทัดจะตัด 'กลางคำ' ตรงไหนบ้าง
    คืน list ของ (ลำดับบรรทัด, คำที่ถูกผ่า, ซ้าย, ขวา) — ไม่แก้ข้อความใด ๆ"""
    if not has_thai(text) or width <= 0:
        return []
    toks = segment(text, engine)
    bad, line_w, line_no = [], 0.0, 1
    for t in toks:
        w = display_width(t)
        if line_w + w > width:
            # คำนี้ล้นบรรทัด — ถ้ามันเป็นคำไทยยาวและไม่มีที่ว่างพอ engine จะผ่ากลางคำ
            if line_w > 0 and has_thai(t) and w > 1:
                room = width - line_w
                if 0 < room < w:
                    cut = int(room)
                    bad.append((line_no, t, t[:cut], t[cut:]))
            line_no += 1
            line_w = w
        else:
            line_w += w
    return bad


AUDIT_SHOW_MAX = 12   # เพดานรายการที่พิมพ์เมื่อไม่ใส่ --all — จำนวนจริงอ่านจากบรรทัดสรุปเสมอ


def audit_xlsx(path: str, col: str = None, width: float = None, show_all: bool = False):
    """T2 กับไฟล์จริง — รายงานเซลล์ที่เสี่ยงตัดกลางคำ (read-only)
    show_all=False พิมพ์ AUDIT_SHOW_MAX รายการแรก + บรรทัดแจ้งยอดที่เหลือ (ห้ามตัดเงียบ)"""
    from openpyxl import load_workbook
    wb = load_workbook(path)
    total, risky = 0, []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if not has_thai(c.value):
                    continue
                if col and c.column_letter != col:
                    continue
                if not (c.alignment and c.alignment.wrap_text):
                    continue
                w = width or (ws.column_dimensions[c.column_letter].width or 20)
                total += 1
                bad = find_bad_breaks(c.value, w)
                if bad:
                    # เก็บตัวอย่างสูงสุด 2 จุดต่อเซลล์ — จำนวนเซลล์เสี่ยงนับจาก len(risky) ไม่ใช่จำนวนบรรทัดที่พิมพ์
                    risky.append((f"{ws.title}!{c.coordinate}", bad[:2]))
    print(f"THAI-WRAP AUDIT | เซลล์ไทยที่ wrap: {total} · เสี่ยงตัดกลางคำ: {len(risky)}")
    shown = risky if show_all else risky[:AUDIT_SHOW_MAX]
    for loc, bad in shown:
        for _, word, l, r in bad:
            print(f"   {loc}: '{word}' → จะถูกผ่าเป็น '{l}' / '{r}'")
    if len(risky) > len(shown):
        print(f"   ⚠ แสดง {len(shown)} จาก {len(risky)} เซลล์เสี่ยง — เหลืออีก "
              f"{len(risky) - len(shown)} เซลล์ ใช้ --all เพื่อดูครบทุกรายการ")
    if risky:
        print("   วิธีแก้ (เรียงจากดีไปรอง): ① ขยายความกว้างคอลัมน์  ② ปรับข้อความ"
              "  ③ ใส่ ZWSP (--zwsp · แลกกับ Ctrl+F หาไม่เจอ)")
    return risky


if __name__ == "__main__":
    a = sys.argv[1:]
    if not a:
        print(__doc__); sys.exit(0)
    if a[0] == "--zwsp":
        print(zwsp(a[1]))
    elif a[0] == "--strip":
        print(strip_zwsp(a[1]))
    elif a[0] == "--check":
        w = float(a[a.index("--width") + 1]) if "--width" in a else 30
        toks = segment(a[1])
        print("ตัดคำได้:", "|".join(toks))
        bad = find_bad_breaks(a[1], w)
        if bad:
            print(f"⚠ ที่ความกว้าง {w} จะตัดกลางคำ:")
            for ln, word, l, r in bad:
                print(f"   บรรทัด {ln}: '{word}' → '{l}' / '{r}'")
        else:
            print(f"✅ ที่ความกว้าง {w} ไม่ตัดกลางคำ")
    elif a[0] == "--audit":
        col = a[a.index("--col") + 1] if "--col" in a else None
        w = float(a[a.index("--width") + 1]) if "--width" in a else None
        audit_xlsx(a[1], col, w, show_all="--all" in a)
    else:
        print(__doc__)
