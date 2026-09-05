#!/usr/bin/env python3
"""
build_xlsx.py — XLSX builder ที่บังคับ FONT POLICY + Thai typography discipline อัตโนมัติ
V02R06 | 2026.08.04
V02R06 — infer_rail: ไม่มี rail ใน spec → เดาจากชนิดเอกสาร + ประกาศเหตุผล
V02R05 — resolve_font_policy ก่อน build (auto-correct) + เลิกส่ง {FONT} เข้า self-audit
V02R04 — V1/V2/V4/V5 เรียก check_fonts จาก SSOT เลิกมีสำเนากฎ | ผูกกับ skill ice-doc-builder §3.0 FONT POLICY + §3.2 E1-E6 + §6 V1-V4

V02R03 — นโยบายย้ายไป `font_policy.py` (SSOT) ไฟล์นี้ import อย่างเดียว ไม่ประกาศซ้ำ
  เหตุผล: RAILS อยู่ในไฟล์นี้ตัวเดียว → build script อีก 5 ตัวใช้ร่วมไม่ได้ → hard-code กันเอง
  ตรวจฟอนต์ทุกฟอร์แมตในที่เดียว → `audit_fonts.py`

V02R02 — ⭐ +V4 RAIL CONFORMANCE (ด่านที่หายไป): V1 ถามแค่ "ชื่อฟอนต์มีจริงไหม" · V2 ถามแค่
"อยู่ blacklist ไหม" → ฟอนต์ที่ **มีจริง + ไม่ blacklist + แต่ผิดนโยบาย** (เช่น Sarabun) ลอดทั้งสองด่าน
เคสจริง 2026.08.04: PWA TCO-Breakdown V01R22 build วันเดียวกับที่นโยบายใช้อยู่แล้ว ยังเป็น Sarabun
และ validator ขึ้น ✅ PASS · ต้นเหตุ: build script เขียนมือตั้ง FONT เองเป็นค่าคงที่ ไม่เคยเรียก RAILS
→ V4 เทียบกับ RAILS[rail] ตรง ๆ · CLI: --rail private|govt · --allow-font "ชื่อ" (TOR บังคับ/ไฟล์รับมา)

V02R03 (2026.08.30) — §3.2 E2 เปลี่ยนจาก "ต้องตั้งความสูงเอง" เป็น "ความสูงต้องพอกับข้อความจริง":
  ค่าเริ่มต้นของ builder = ปล่อย auto-height · ตั้งเองเฉพาะแถว merge หรือแถวที่ spec สั่ง row_height_lines
  audit เลิก fail แถวที่ปล่อย auto · หันมา fail แถวที่ตั้งเองแล้วเตี้ยกว่าที่ข้อความต้องการ
  (คำนวณจากความกว้างคอลัมน์จริงของเซลล์นั้น ไม่ใช่การเดาจำนวนบรรทัดแบบเหมารวม)
  หลักฐาน: CP Axtra Requirement Baseline 2026.08.30 — ×1.45 เฉือน 83/112 แถว · ×1.72 เฉือน 52/140 · auto 0/140

V02R02 — แก้ E4 false positive: เดิมฟ้อง "merge ในแถวไทย+wrap" ทุกแถวโดยไม่ดูว่าแถวนั้นตั้ง
row height ไว้แล้วหรือไม่ → ไฟล์ที่ builder ของเราสร้างสด (merge_put ตั้ง height ให้เสมอ) ก็ FAIL
พิสูจน์ด้วย differential test: rebuild จาก script เดิม → FAIL เซลล์ชุดเดียวกันเป๊ะ = ปัญหาอยู่ที่
เครื่องมือตรวจ ไม่ใช่ไฟล์ (หลักการเดียวกับบทเรียน renderer shim 2026.08.01)

ทำไมต้องมีชั้นนี้: กฎที่เป็นตัวหนังสือใน skill ถูกลืมได้ · เคสจริง 2026.07.31 ไฟล์ PWA TOR Matrix
ระบุฟอนต์ "IBM Plex Sans Thai Regular" (ไม่มี family ชื่อนี้จริง) → Excel substitute เงียบ →
ฟอนต์ปน 3 ตัวในไฟล์เดียว · ไม่มี validator ตัวใดในระบบจับได้ → ตอนนี้ V1 จับได้ตั้งแต่ก่อน save

Usage:
    python3 build_xlsx.py spec.json out.xlsx        # ต้องมี marker ICE_BUILD=pipeline นำหน้า
    python3 build_xlsx.py --audit existing.xlsx     # ตรวจไฟล์ที่มีอยู่แล้ว (ไม่เขียนอะไร)

spec.json schema (เพิ่มจาก V01 · ของเดิมใช้ได้เหมือนเดิมทุกอย่าง):
{
  "rail": "private" | "govt",          # เลือกราง (default: private) → กำหนดฟอนต์+ขนาดตาม §3.0
  "font": "…", "font_size": 11,        # override ราง (ใช้เมื่อ TOR/ลูกค้าบังคับ)
  "sheets": [{
      "name": "Sheet1",
      "headers": [...], "rows": [[...]],
      "formulas": {"D2": "=B2+C2"},    # ⚠ ระวัง §2B.3 Removed Records
      "freeze": "B2",
      "column_widths": {"A": 30},
      "header_fill": "1F4E79", "header_font_color": "FFFFFF",
      "wrap_columns": ["C","G"],       # คอลัมน์ที่ให้ wrap — V02R03: ปล่อย auto-height ให้โปรแกรมคำนวณเอง
      "row_height_lines": {"5": 2}     # สั่งความสูงคงที่ของแถวนั้น (= เลิกใช้ auto สำหรับแถวนี้)
                                       #   ใช้เมื่อดีไซน์ต้องการความสูงตายตัว เช่น banner — ดู §3.2 E2
  }]
}
"""
import sys, os, json, glob, re, math

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─────────────────────────────────────────────────────────────────────────────
# §3.0 FONT POLICY — ⭐ import จาก SSOT (V02R03) · ห้าม re-declare ที่ไฟล์นี้อีก
#   บทเรียน 2026.08.04: RAILS เคยถูกประกาศซ้ำในไฟล์นี้ → build script ฟอร์แมตอื่น
#   (pptx/docx/dashboard/deck/html) ไม่มีทางใช้ร่วมได้ → hard-code กันเอง 5 ระบบ
#   → นโยบายที่ LOCKED ไว้ บังคับใช้ไม่ได้จริงทั้งระบบ (เคส PWA TCO V01R22)
# ─────────────────────────────────────────────────────────────────────────────
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from font_policy import (RAILS, BLACKLIST_PATTERNS, LATIN_ONLY, THAI_RE,   # noqa: E402
                         APPROVED_ALT, RETIRED, rail_fallbacks, resolve_font_policy, infer_rail,
                         has_thai, installed_families, blacklist_hit, check_fonts)

# §3.2 E2 (V02R03 · 2026.08.30) — ค่าเหล่านี้เป็น "พื้นขั้นต่ำ" สำหรับแถวที่ *จำเป็น* ต้องตั้งความสูงเอง
# (แถว merge ที่ AutoFit ตาย และแถว banner ที่ดีไซน์กำหนดความสูง) ไม่ใช่ค่าที่ใช้กับทุกแถวอีกต่อไป
# ค่าเริ่มต้นของแถวข้อความยาว = ปล่อย auto-height ให้โปรแกรมคำนวณ
# ที่มา: CP Axtra Requirement Baseline 2026.08.30 — ×1.45 เฉือน 83/112 แถว · ×1.72 เฉือน 52/140 · auto 0/140
ROW_H_FACTOR = 1.72   # ตัวคูณตอน "สร้าง" — เผื่อไว้เล็กน้อย (เดิม 1.45 ต่ำเกินไปจนเฉือนข้อความ)
# ตัวคูณตอน "ตรวจ" ต้องต่ำกว่าตัวคูณตอนสร้าง มิฉะนั้นจะฟ้องแถวที่ปกติดี
#   วัดความสูงหนึ่งบรรทัดจริงด้วย LibreOffice (2026.08.30): IBM Plex Sans Thai Looped = 1.67-1.83 × pt
#   · TH Sarabun New = 1.40-1.49 × pt → ตั้งพื้นที่ 1.35 ซึ่งต่ำกว่าทุกค่าที่วัดได้ จึงฟ้องเฉพาะแถวที่สั้นผิดปกติจริง
#   ส่วนคำถามว่า "ถูกเฉือนหรือไม่" ให้ xlsx_rowheight_probe.py ตอบ (§3.2 E2 ชั้น ②) ไม่ใช่ค่านี้
ROW_H_MIN_FACTOR = 1.35   # ใช้เมื่ออ่าน metric ของฟอนต์ไม่ได้เท่านั้น (ดู min_single_line_height)
ROW_H_PAD = 6
ROW_H_TOL = 2.0       # ผ่อนผัน 2 pt กันเสียงรบกวนจากการประมาณ — เฉือนจริงในเคสอ้างอิงคือ 11-21 pt


def min_single_line_height(font_pt):
    """พื้นขั้นต่ำของแถวหนึ่งบรรทัด (point) แบบไม่รู้จักฟอนต์ — **ใช้เมื่ออ่าน metric ไม่ได้เท่านั้น**
    ทางหลักคือ `line_ratio()` ซึ่งอ่านจากไฟล์ฟอนต์จริง ค่าคงที่ 1.35 นี้เลือกให้ต่ำกว่าทุกฟอนต์ที่วัดได้

    ⚠ ข้อจำกัดที่ต้องรู้ และเป็นเหตุผลที่ฟังก์ชันนี้ไม่พยายามทำมากกว่านี้:
      การพยากรณ์ว่าข้อความไทยจะ wrap กี่บรรทัด ทำจากข้อมูลสถิตไม่ได้จริง — วัดแล้วสองทาง
      (หลักฐาน CP Axtra 2026.08.30 · เทียบกับผลวัด render จริง 52 แถว):
        · นับทุกอักขระ ÷ 1.18            → ฟ้อง 105 แถว (เกินจริงราวสองเท่า = validator ที่เชื่อไม่ได้)
        · ตัดสระ/วรรณยุกต์ที่กว้าง 0 แล้วใช้อัตราส่วนที่วัดจากไฟล์ฟอนต์จริง (1.058) → ฟ้อง 7 แถว (ต่ำกว่าจริงมาก)
      เพราะหน่วยความกว้างคอลัมน์ของ Excel ไม่ใช่จำนวนตัวอักษรตรง ๆ และ LibreOffice ตัดบรรทัดไทย
      ด้วยพจนานุกรม ไม่ใช่ตามจำนวนอักขระ
      ⇒ ตัวตรวจสถิตจึงตอบได้แค่ "เตี้ยจนไม่พอแม้บรรทัดเดียวหรือไม่" ส่วน "ถูกเฉือนหรือไม่"
        ต้องให้เครื่องมือจัดหน้าตอบ — ใช้ `xlsx_rowheight_probe.py` (§3.2 E2 ชั้น ②)
        การแกล้งตอบให้ดูขยันคือการสร้าง false-green รอบใหม่
    """
    return round(font_pt * ROW_H_MIN_FACTOR, 1)


_LINE_RATIO_CACHE = {}


def line_ratio(family):
    """สัดส่วนความสูงหนึ่งบรรทัดต่อขนาดฟอนต์ อ่านจากไฟล์ฟอนต์จริง (hhea) — คืน None ถ้าอ่านไม่ได้

    ทำไมต้องอ่านจากไฟล์ ไม่ใช้ค่าคงที่: ค่าที่ถูกของฟอนต์หนึ่งผิดสำหรับอีกฟอนต์
      IBM Plex Sans Thai Looped = 1.650 · TH Sarabun New = 1.331 — ต่างกัน 24%
    ทำไมใช้ hhea ไม่ใช้ usWinAscent+usWinDescent: วัดเทียบกับความสูงที่ LibreOffice กำหนดจริง
      (2026.08.30 · 6 เคส) แล้วพบว่าฐาน usWin สูงเกินจนฟ้องผิด 3 ใน 6 เคส ส่วนฐาน hhea
      อยู่ใต้ค่าจริงทุกเคส จึงใช้เป็น "พื้น" ได้โดยไม่ฟ้องแถวที่ปกติดี
      (Plex 9pt: hhea 14.8 · จริง 15.0 | 16pt: hhea 26.4 · จริง 26.9 | Sarabun 16pt: hhea 21.3 · จริง 22.4)
    """
    if family in _LINE_RATIO_CACHE:
        return _LINE_RATIO_CACHE[family]
    ratio = None
    try:
        from font_policy import font_file_for
        from fontTools.ttLib import TTFont
        path = font_file_for(family)
        if path:
            f = TTFont(path, fontNumber=0, lazy=True)
            h, upm = f["hhea"], f["head"].unitsPerEm
            ratio = (h.ascender - h.descender + h.lineGap) / upm
    except Exception:
        ratio = None
    _LINE_RATIO_CACHE[family] = ratio
    return ratio


# ─────────────────────────────────────────────────────────────────────────────
# VALIDATOR — §6 V1/V2 + §3.2 E2-E5
# ─────────────────────────────────────────────────────────────────────────────
def audit(path: str, strict: bool = True, rail: str = "private",
          allow_fonts=None) -> dict:   # ⚠ ไม่ใช้ `set | None` — python3 บนเครื่องนี้คือ 3.9
    """rail = private|govt (ตาม §3.0) · allow_fonts = ฟอนต์ที่ยกเว้นได้ (TOR บังคับ / ไฟล์ที่ได้รับมา)"""
    allow_fonts = allow_fonts or set()
    wb = load_workbook(path)
    fams = installed_families()
    rep = {"fonts_used": set(), "unresolvable": [], "blacklisted": [],
           "thai_cells": 0, "thai_latin_only_font": [],
           "row_too_short": [], "row_auto_height": 0, "row_fixed_height": 0,
           "merged_thai_rows": [], "merged_thai_rows_ok": [],
           "shrink_to_fit": [], "formula_cells": 0,
           "v1_skipped": not bool(fams)}

    for ws in wb.worksheets:
        merged_rows = set()
        for rng in ws.merged_cells.ranges:
            merged_rows.update(range(rng.min_row, rng.max_row + 1))
        for row in ws.iter_rows():
            for c in row:
                if c.value is None:
                    continue
                fname = c.font.name if c.font else None
                if fname:
                    rep["fonts_used"].add(fname)
                if isinstance(c.value, str) and c.value.startswith("="):
                    rep["formula_cells"] += 1
                if not has_thai(c.value):
                    continue
                rep["thai_cells"] += 1
                loc = f"{ws.title}!{c.coordinate}"
                if not fname or fname in LATIN_ONLY:
                    rep["thai_latin_only_font"].append((loc, fname or "(ไม่ระบุ)"))
                if c.alignment and c.alignment.shrinkToFit:
                    rep["shrink_to_fit"].append(loc)
                wrapped = bool(c.alignment and c.alignment.wrap_text)
                if wrapped:
                    set_h = ws.row_dimensions[c.row].height
                    has_h = set_h is not None
                    if not has_h:
                        # V02R03: ปล่อย auto-height = ค่าเริ่มต้นที่ถูกต้องตาม E2 ใหม่ ไม่ใช่ข้อบกพร่อง
                        #   (E4 ด้านล่างตัดสินแถว merge จากตัวแปร has_h ในเครื่อง ไม่ได้ใช้คีย์ใด)
                        rep["row_auto_height"] += 1
                    else:
                        # ตั้งเอง = ตรวจได้เฉพาะข้อที่ฟันธงได้: เตี้ยจนไม่พอแม้บรรทัดเดียวหรือไม่
                        rep["row_fixed_height"] += 1
                        pt = float((c.font.sz if (c.font and c.font.sz) else 11) or 11)
                        r = line_ratio(c.font.name if c.font else None)
                        base = round(pt * r, 1) if r else min_single_line_height(pt)
                        # นับบรรทัดจากการขึ้นบรรทัดใหม่ในเซลล์ — รู้แน่ ไม่ต้องพยากรณ์ wrap
                        #   (แถวที่บรรจุ 3 บรรทัดแต่ตั้งความสูงไว้เท่าบรรทัดเดียว เดิมลอดด่านนี้ไปได้)
                        floor = round(base * max(1, len(str(c.value).splitlines())), 1)
                        if set_h < floor - ROW_H_TOL:
                            rep["row_too_short"].append((loc, round(set_h, 1), floor))
                    # E4 — merge ฆ่า AutoFit จริง แต่เป็น "ข้อบกพร่อง" ก็ต่อเมื่อแถวนั้น *ต้องพึ่ง* AutoFit
                    #   ตั้ง row height ไว้ชัดเจนแล้ว = escape hatch ที่ builder เราใช้ตั้งใจ (merge_put) → ผ่าน
                    #   V02R02 (บทเรียน PWA 2026.08.04): เดิมไม่เช็คเงื่อนไขนี้ → ไฟล์ที่ script สร้างสด
                    #   และผ่าน assert ตัวเอง ก็ยัง FAIL E4 = false positive ทุกแถว banner/band
                    if c.row in merged_rows:
                        (rep["merged_thai_rows"] if not has_h
                         else rep["merged_thai_rows_ok"]).append(loc)

    for name in sorted(rep["fonts_used"]):
        why = blacklist_hit(name)
        if why:
            rep["blacklisted"].append((name, why))
        if fams and name not in fams and name not in LATIN_ONLY:
            rep["unresolvable"].append(name)

    # §3.5 T2 — ทำนายจุดตัดกลางคำ (เตือน ไม่ fail: แก้ที่ความกว้างคอลัมน์ ไม่ใช่ที่ข้อความ)
    try:
        sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
        from thai_wordbreak import find_bad_breaks
        risky = 0
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    if has_thai(c.value) and c.alignment and c.alignment.wrap_text:
                        w = ws.column_dimensions[c.column_letter].width or 20
                        if find_bad_breaks(c.value, w):
                            risky += 1
        rep["thai_midword_break"] = risky
        if risky:
            print(f"⚠ §3.5 T2: {risky} เซลล์เสี่ยงตัดบรรทัดกลางคำ → "
                  f"`python3 _lib/thai_wordbreak.py --audit <file>` ดูรายจุด "
                  f"(แก้: ①ขยายคอลัมน์ ②ปรับข้อความ ③ZWSP-แลกกับ Ctrl+F)")
    except Exception:
        pass

    # ⭐ V02R04: V1/V2/V4/V5 เรียก check_fonts จาก SSOT — เลิกมีสำเนากฎในไฟล์นี้
    #   (เดิม V02R02 เขียน logic ซ้ำที่นี่ = ตรงข้ามกับเหตุผลที่แยก font_policy ออกมาแต่แรก)
    fc = check_fonts(rep["fonts_used"], rail=rail, allow_fonts=allow_fonts, fams=fams)
    for k in ("off_rail", "retired", "alt_used", "rail"):
        rep[k] = fc[k]
    fails = list(fc["fails"])

    if rep["thai_latin_only_font"]:
        fails.append(f"เซลล์ไทยได้ฟอนต์ที่ไม่มี glyph ไทย: {rep['thai_latin_only_font'][:5]}")
    if rep["merged_thai_rows"]:
        fails.append(f"E4 merge ในแถวไทย+wrap *และไม่ตั้ง row height* (AutoFit ตาย ไม่มีอะไรกู้ได้): "
                     f"{rep['merged_thai_rows'][:5]}")
    if rep["shrink_to_fit"]:
        fails.append(f"E5 shrink-to-fit บนเซลล์ไทย: {rep['shrink_to_fit'][:5]}")
    if rep["row_too_short"]:
        _ex = "; ".join(f"{loc} ตั้งไว้ {h} pt · พื้นขั้นต่ำ {n} pt" for loc, h, n in rep["row_too_short"][:5])
        fails.append(f"E2 แถวที่ตั้งความสูงเองแล้วไม่พอกับจำนวนบรรทัดที่มีอยู่จริง "
                     f"({len(rep['row_too_short'])} เซลล์ · นับจากบรรทัดที่ขึ้นใหม่ในเซลล์ ไม่ใช่การเดา wrap) — ขยายให้ถึงค่าที่ระบุ "
                     f"หรือเปลี่ยนมาปล่อย auto-height: {_ex}")
    rep["fails"] = fails

    print(f"VALIDATOR | thai_cells={rep['thai_cells']} · fonts={len(rep['fonts_used'])} "
          f"· formulas={rep['formula_cells']} · V1={'SKIPPED(no fontTools)' if rep['v1_skipped'] else 'ran'}")
    print(f"  fonts: {', '.join(sorted(rep['fonts_used'])) or '(none)'}")
    if rep.get("alt_used"):
        print(f"  ℹ ใช้ตัวเลือกอนุมัติ (ไม่ใช่ฟอนต์ราง): {', '.join(rep['alt_used'])} "
              f"— ผ่านได้ แต่ GAP ไทย-ละตินกว้างกว่า ตรวจว่าไทยไม่ดูเล็กเกินไป")
    if rep["row_auto_height"] or rep["row_fixed_height"]:
        # แสดงเสมอ — ผู้อ่านต้องรู้ว่าอะไร "ผ่านเพราะตรวจแล้ว" กับอะไร "ผ่านเพราะตรวจไม่ได้"
        print(f"  ℹ E2 ความสูงแถว: {rep['row_auto_height']} เซลล์ปล่อย auto-height (ค่าเริ่มต้นตาม §3.2 E2) · "
              f"{rep['row_fixed_height']} เซลล์ตั้งความสูงเอง (ตรวจแล้วว่าไม่ต่ำกว่าพื้นหนึ่งบรรทัด)")
        print(f"     ⚠ ตัวตรวจนี้บอกไม่ได้ว่าข้อความถูกเฉือนหรือไม่ — การ wrap ภาษาไทยพยากรณ์จากข้อมูลสถิตไม่ได้ "
              f"(ดูเหตุผลใน min_single_line_height) → ตอบข้อนั้นด้วย xlsx_rowheight_probe.py")
    if rep["merged_thai_rows_ok"]:
        # แสดงเสมอ — ไม่เงียบ: ผู้อ่านต้องรู้ว่าเรา "ยกเว้น" อะไรไป ไม่ใช่ "ไม่เจอ"
        print(f"  ℹ E4 ยกเว้น {len(rep['merged_thai_rows_ok'])} เซลล์ merge+wrap ที่ตั้ง row height ไว้แล้ว "
              f"(escape hatch ถูกกฎ) เช่น {rep['merged_thai_rows_ok'][:3]}")
    if fails:
        print("❌ FAIL:")
        for f in fails:
            print("   -", f)
        if strict:
            sys.exit(3)
    else:
        v1txt = "V1 ข้าม (ไม่มี fontTools — ข้าม ≠ ผ่าน)" if rep["v1_skipped"] else "V1 resolve ครบ"
        print(f"✅ PASS — {v1txt} · V2 ไม่มี blacklist · V4 ตรงราง '{rail}' · E2/E4/E5 ผ่าน "
              f"(E2 = ไม่มีแถวที่ตั้งเองต่ำกว่าพื้นหนึ่งบรรทัด · **ไม่ได้ตรวจ**การเฉือนจาก wrap)")
    return rep


# ─────────────────────────────────────────────────────────────────────────────
# BUILD
# ─────────────────────────────────────────────────────────────────────────────
def build(spec_path: str, out_path: str):
    with open(spec_path) as f:
        spec = json.load(f)

    # ⭐ V02R06: ไม่มี rail ใน spec → เดาจากชนิดเอกสาร (path/ชื่อไฟล์/หัวเรื่อง) + ประกาศเสมอ
    rail, _why = infer_rail(spec, out_path)
    print(f"📄 ราง: {rail}  ({_why})")
    if rail not in RAILS:
        sys.exit(f"rail ต้องเป็น private|govt (ได้: {rail})")
    FONT = spec.get("font", RAILS[rail]["font"])
    SIZE = spec.get("font_size", RAILS[rail]["size"])

    # V1/V2 ตรวจ "ก่อน" build — ล้มเร็วดีกว่าล้มช้า
    why = blacklist_hit(FONT)
    if why:
        sys.exit(f"❌ V2 BLACKLIST: '{FONT}' — {why}")
    fams = installed_families()
    # ⭐⭐ V02R05 (2026.08.04) — ด่านนโยบาย **ก่อน build** · แก้ให้อัตโนมัติ ไม่ fail
    #   บทเรียน VFIN: spec ระบุ Sarabun → build ผ่าน → self-audit PASS (เพราะถูกยื่นเฉลย
    #   `allow_fonts={FONT}` = self-approving checker) → เจอ FAIL ตอน audit แยก → เสีย 1 รอบเต็ม
    #   ⇒ ย้ายการตัดสินมาก่อนเขียนไฟล์ + **แก้ให้เลยแทนที่จะ fail** (คำตอบถูกมีตัวเดียว = ฟอนต์ราง)
    FONT, _notices, _allow = resolve_font_policy(FONT, rail, spec, fams)
    for _n in _notices:
        print(_n)

    wb = Workbook()
    wb.remove(wb.active)

    # §3.2 E6 — ตั้ง default font ก่อน แล้วค่อยคำนวณ column width
    for ns in wb._named_styles:
        if ns.name == "Normal":
            ns.font = Font(name=FONT, size=SIZE)

    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for sht in spec.get("sheets", []):
        ws = wb.create_sheet(title=sht["name"][:31])
        wrap_cols = set(sht.get("wrap_columns") or [])
        headers = sht.get("headers", [])
        if headers:
            ws.append(headers)
            hf = PatternFill("solid", fgColor=sht.get("header_fill", "1F4E79"))
            hfont = Font(name=FONT, size=SIZE, bold=True,
                         color=sht.get("header_font_color", "FFFFFF"))
            for j, _ in enumerate(headers, start=1):
                c = ws.cell(row=1, column=j)
                c.fill, c.font = hf, hfont
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in sht.get("rows", []):
            ws.append(row)
        for cell, formula in (sht.get("formulas") or {}).items():
            ws[cell] = formula

        overrides = {int(k): v for k, v in (sht.get("row_height_lines") or {}).items()}

        for row in ws.iter_rows():
            thai_in_row = False
            wrapped_in_row = False
            for c in row:
                col = c.column_letter
                is_thai = has_thai(c.value)
                thai_in_row = thai_in_row or is_thai
                # ⭐ ฟอนต์เดียวทุกเซลล์ (§3.0 single-family) — กัน default leak
                keep_bold = bool(c.font and c.font.bold)
                keep_color = c.font.color if (c.font and c.font.color) else None
                c.font = Font(name=FONT, size=SIZE, bold=keep_bold, color=keep_color)
                c.border = border
                want_wrap = col in wrap_cols
                wrapped_in_row = wrapped_in_row or want_wrap
                # §3.2 E3 vertical=center · E5 ไม่ตั้ง shrinkToFit เด็ดขาด
                c.alignment = Alignment(
                    horizontal=(c.alignment.horizontal if c.alignment else None),
                    vertical="center", wrap_text=want_wrap or (c.alignment.wrap_text if c.alignment else False),
                    shrink_to_fit=False,
                )
            # §3.2 E2 (V02R03) — ค่าเริ่มต้นคือ **ปล่อย auto-height**
            #   ตั้งเองเฉพาะ 2 กรณีที่ auto ใช้ไม่ได้หรือดีไซน์บังคับ:
            #     (ก) แถวที่ merge — AutoFit ถูกปิดใช้งานที่นั่น (E4) ไม่ตั้ง = ไม่มีอะไรกู้
            #     (ข) spec สั่งจำนวนบรรทัดมาเอง (row_height_lines) = ผู้เขียน spec ยืนยันความสูงคงที่
            #   เหตุผลที่เลิกตั้งทุกแถว: การเดา "2 บรรทัดถ้ายาวเกิน 60 ตัว" ผิดบ่อยจนเฉือนข้อความ
            #   (CP Axtra 2026.08.30 — เฉือน 52/140 แถวแม้ใช้ตัวคูณ 1.72 แล้ว · ปล่อย auto = 0/140)
            rownum = row[0].row
            spec_lines = overrides.get(rownum)
            # จำนวนบรรทัดที่ "รู้แน่" จากการขึ้นบรรทัดใหม่ในเซลล์ — ไม่ใช่การพยากรณ์ wrap จึงเชื่อได้ 100%
            hard_lines = max((len(str(c.value).splitlines()) for c in row if c.value), default=1)
            if spec_lines is not None or hard_lines > 1:
                lines = max(spec_lines or 1, hard_lines)
                ws.row_dimensions[rownum].height = round(SIZE * ROW_H_FACTOR * lines + ROW_H_PAD, 1)
            # นอกเหนือจากนี้ = ปล่อย auto-height ตาม §3.2 E2
            #   (เดิมมีสาขาสำหรับแถว merge ด้วย แต่ builder นี้ไม่เคยเรียก merge_cells และ spec ก็ไม่มีคีย์
            #    สำหรับสั่ง merge → เป็น dead code ที่ซ่อนสูตรเดา "2 บรรทัดถ้ายาวเกิน 60 ตัว" ซึ่งเลิกใช้แล้วไว้ข้างใน
            #    ถ้าวันหนึ่งเพิ่ม merge เข้ามาจริง ต้องเพิ่มคีย์ใน schema แล้วตั้งความสูงให้แถวนั้นด้วย เพราะ E4)

        for col, width in (sht.get("column_widths") or {}).items():
            ws.column_dimensions[col].width = width
        if sht.get("freeze"):
            ws.freeze_panes = sht["freeze"]

    wb.save(out_path)
    if not os.path.exists(out_path):          # กฎเหล็ก: tool สำเร็จ ≠ ไฟล์เกิด
        sys.exit("❌ save รายงานสำเร็จแต่ไม่พบไฟล์")
    print(f"OK: {out_path} · {len(wb.sheetnames)} sheets · rail={rail} · font={FONT} {SIZE}pt")
    print(f"⚠ §3.2 E1: Excel ฝังฟอนต์ไม่ได้ — ส่งลูกค้าต้องแนบ PDF companion หรือใช้ฟอนต์ที่ลูกค้ามี "
          f"(fallback ตามลำดับ: {' → '.join(rail_fallbacks(rail))})")
    # ⭐ V02R05: ส่ง **เฉพาะ** ที่ spec ประกาศเหตุผลไว้ — ห้ามส่ง {FONT} (= ยื่นเฉลยให้ผู้ตรวจ)
    #   rail ที่ถูกต้องคือสิ่งที่กัน build รางราชการ FAIL กับ default private — ไม่ใช่ whitelist ตัวเอง
    audit(out_path, rail=rail, allow_fonts=_allow)


if __name__ == "__main__":
    argv = sys.argv[1:]
    rail, allow = "private", set()
    while len(argv) >= 2 and argv[0] in ("--rail", "--allow-font"):
        if argv[0] == "--rail":
            rail = argv[1]
            if rail not in RAILS:
                sys.exit(f"--rail ต้องเป็น private|govt (ได้: {rail})")
        else:
            allow.add(argv[1])
        argv = argv[2:]

    if len(argv) == 2 and argv[0] == "--audit":
        audit(argv[1], rail=rail, allow_fonts=allow); sys.exit(0)
    if len(argv) != 2:
        print("usage: build_xlsx.py [--rail private|govt] [--allow-font NAME]... "
              "(spec.json out.xlsx | --audit file.xlsx)", file=sys.stderr)
        sys.exit(2)
    build(argv[0], argv[1])
