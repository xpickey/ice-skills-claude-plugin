#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""xlsx_rowheight_probe.py — ตรวจว่า "ความสูงแถวที่ตั้งไว้ พอกับข้อความจริงหรือไม่"
V01R01 | 2026.08.30 | ผูกกับ skill ice-doc-builder §3.2 E2 ชั้น ②

ทำไมต้องมีเครื่องมือนี้ (อย่าลบทิ้งแล้วกลับไปใช้สูตร):
  §3.2 E2 บังคับให้ตัดสินที่ผลลัพธ์ ไม่ใช่ที่วิธีตั้งค่า แต่ตัวตรวจสถิตทำแทนไม่ได้ —
  ลองแล้วสองทางกับไฟล์จริงที่เฉือน 52 แถว (CP Axtra 2026.08.30):
    · นับอักขระ ÷ 1.18                                  → ฟ้อง 105 แถว (เกินจริงราวสองเท่า)
    · ตัดสระ/วรรณยุกต์ที่กว้าง 0 + อัตราส่วนจากไฟล์ฟอนต์จริง → ฟ้อง 7 แถว (ต่ำกว่าจริงมาก)
  เพราะหน่วยความกว้างคอลัมน์ของ Excel ไม่ใช่จำนวนตัวอักษร และการตัดบรรทัดไทยใช้พจนานุกรม
  ⇒ วิธีเดียวที่เชื่อได้คือ **ถามเครื่องมือจัดหน้าเองว่าแถวนี้ต้องสูงเท่าไร**

วิธีทำงาน (3 ขั้น):
  ① คัดลอกไฟล์แล้ว **ล้างความสูงแถวออกทั้งหมด** ลงไฟล์ชั่วคราว
  ② ให้ LibreOffice ตัวจริงแปลงไฟล์นั้นเป็น .xlsx อีกครั้ง — ระหว่างแปลงมันคำนวณ auto-height
     ให้ทุกแถวแล้วเขียนค่าลงไฟล์ ค่านั้นคือ "ความสูงที่ข้อความต้องการจริง"
  ③ เทียบกับความสูงที่ไฟล์ต้นฉบับตั้งไว้ — ต่ำกว่า = ข้อความถูกเฉือน

ข้อจำกัดที่ต้องรู้:
  · ใช้ LibreOffice เป็นผู้ตัดสิน ไม่ใช่ Excel — ค่าอาจต่างจาก Excel เล็กน้อย แต่เป็นหลักฐานที่วัดได้
    และดีกว่าการเดาจากจำนวนอักขระอย่างเทียบไม่ติด
  · แถวที่ไม่ได้ตั้งความสูง (auto) จะข้ามไป เพราะไม่มีอะไรให้เทียบ — แถวพวกนั้นถูกต้องโดยโครงสร้างอยู่แล้ว
  · ต้องมี LibreOffice ตัวจริง (ไม่ใช่ shim ใน PATH) — ตรวจด้วย `render_pdf.sh --which` ถ้าหาไม่เจอ

Usage:
    python3 xlsx_rowheight_probe.py FILE.xlsx [--tol 1.0] [--quiet]
Exit: 0 = ทุกแถวพอ · 3 = มีแถวถูกเฉือน · 2 = ใช้งานไม่ได้ (ไม่มีไฟล์ / ไม่มี LibreOffice)
"""
import os
import shutil
import subprocess
import sys
import tempfile

LO_CANDS = [
    "/Applications/LibreOffice.app/Contents/MacOS/soffice",
    os.path.expanduser("~/Applications/LibreOffice.app/Contents/MacOS/soffice"),
    "/usr/lib/libreoffice/program/soffice",
]


def find_real_lo():
    """หา LibreOffice ตัวจริง — ห้ามใช้ soffice จาก PATH เพราะบนเครื่องนี้เป็น shim
    ที่มองไม่เห็นฟอนต์ระบบแล้วแทนฟอนต์เงียบ ๆ (เหตุผลเต็มอยู่ใน render_pdf.sh)"""
    import glob
    cands = list(LO_CANDS) + glob.glob(
        "/opt/homebrew/Caskroom/libreoffice/*/LibreOffice.app/Contents/MacOS/soffice")
    for c in cands:
        if not os.access(c, os.X_OK):
            continue
        try:
            out = subprocess.run([c, "--version"], capture_output=True, text=True, timeout=60).stdout
        except Exception:
            continue
        if out.startswith("LibreOffice "):
            return c
    return None


def probe(path, tol=1.0, quiet=False):
    from openpyxl import load_workbook
    if not os.path.isfile(path):
        print("❌ ไม่พบไฟล์: %s" % path, file=sys.stderr)
        return 2
    lo = find_real_lo()
    if not lo:
        print("❌ ไม่พบ LibreOffice ตัวจริง — ตรวจด้วย `bash ~/.claude/agents/_lib/render_pdf.sh --which`",
              file=sys.stderr)
        print("   (ข้ามการตรวจนี้ ≠ ผ่าน — ต้องหา LibreOffice ให้ได้ก่อนถือว่างานผ่าน E2 ชั้น ②)",
              file=sys.stderr)
        return 2

    original = load_workbook(path)
    tmpdir = tempfile.mkdtemp(prefix="rowheight-probe-")
    try:
        stripped = os.path.join(tmpdir, "stripped.xlsx")
        wb = load_workbook(path)
        for ws in wb.worksheets:
            for r in list(ws.row_dimensions):
                ws.row_dimensions[r].height = None
        wb.save(stripped)

        outdir = os.path.join(tmpdir, "out")
        subprocess.run([lo, "--headless", "-env:UserInstallation=file://%s/profile" % tmpdir,
                        "--convert-to", "xlsx", "--outdir", outdir, stripped],
                       capture_output=True, timeout=600)
        computed_path = os.path.join(outdir, "stripped.xlsx")
        if not os.path.isfile(computed_path):
            print("❌ LibreOffice แปลงไฟล์ไม่สำเร็จ — ตรวจว่าไฟล์เปิดได้จริงก่อน", file=sys.stderr)
            return 2
        computed = load_workbook(computed_path)

        short, compared, auto = [], 0, 0
        for ws in original.worksheets:
            if ws.title not in computed.sheetnames:
                continue
            ref = computed[ws.title]
            for r in range(1, ws.max_row + 1):
                set_h = ws.row_dimensions[r].height
                need_h = ref.row_dimensions[r].height
                if set_h is None:
                    auto += 1
                    continue
                if need_h is None:
                    continue
                compared += 1
                if set_h < need_h - tol:
                    short.append((ws.title, r, round(set_h, 1), round(need_h, 1)))
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)

    if not quiet:
        print("ROW-HEIGHT PROBE | เทียบ %d แถวที่ตั้งความสูงเอง · ข้าม %d แถวที่ปล่อย auto (ถูกต้องโดยโครงสร้าง)"
              % (compared, auto))
    if short:
        print("❌ E2 ชั้น ② — พบ %d แถวที่ความสูงไม่พอกับข้อความ (ข้อความถูกเฉือนจริง):" % len(short))
        for t, r, a, b in short[:15]:
            print("   %s แถว %d : ตั้งไว้ %.1f pt · ต้องการ %.1f pt (ขาด %.1f)" % (t, r, a, b, b - a))
        if len(short) > 15:
            print("   … อีก %d แถว" % (len(short) - 15))
        print("   ทางแก้: ปล่อยแถวเหล่านี้เป็น auto-height (แนะนำ) หรือขยายให้ถึงค่าที่ระบุ แล้วรันซ้ำ")
        return 3
    print("✅ E2 ชั้น ② ผ่าน — ทุกแถวที่ตั้งความสูงเองมีที่พอสำหรับข้อความจริง")
    return 0


if __name__ == "__main__":
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    tol = 1.0
    if "--tol" in sys.argv:
        tol = float(sys.argv[sys.argv.index("--tol") + 1])
    if not args:
        print(__doc__)
        sys.exit(2)
    sys.exit(probe(args[0], tol=tol, quiet="--quiet" in sys.argv))
